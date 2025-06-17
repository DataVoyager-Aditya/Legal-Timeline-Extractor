
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from io import BytesIO
import logging
from typing import List, Dict, Any

class ExcelExporter:
    """
    Professional Excel exporter for legal timeline analysis.
    Creates comprehensive workbooks with multiple sheets and analysis.
    """

    def __init__(self):
        """Initialize Excel exporter"""
        self.logger = logging.getLogger(__name__)

    def create_timeline_workbook(self, timeline_events: List[Dict[str, Any]], metadata: Dict[str, Any]) -> BytesIO:
        """
        Create comprehensive Excel workbook with timeline analysis.

        Args:
            timeline_events: List of timeline events
            metadata: Report metadata

        Returns:
            BytesIO buffer containing Excel data
        """
        buffer = BytesIO()

        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            self._create_summary_sheet(wb, timeline_events, metadata)
            self._create_timeline_sheet(wb, timeline_events)
            self._create_analysis_sheet(wb, timeline_events)
            self._create_evidence_sheet(wb, timeline_events)

            # Save to buffer
            wb.save(buffer)
            buffer.seek(0)

            self.logger.info("Excel workbook generated successfully")
            return buffer

        except Exception as e:
            self.logger.error(f"Error creating Excel workbook: {e}")
            raise

    def _create_summary_sheet(self, wb: Workbook, timeline_events: List[Dict[str, Any]], metadata: Dict[str, Any]):
        """Create summary sheet with overview statistics"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "LEGAL TIMELINE ANALYSIS SUMMARY"
        ws['A1'].font = Font(size=16, bold=True, color="1e3c72")
        ws.merge_cells('A1:E1')

        # Metadata section
        row = 3
        ws[f'A{row}'] = "Report Information"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        info_data = [
            ("Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Total Events:", len(timeline_events)),
            ("Source Documents:", len(metadata.get('source_files', []))),
            ("Analysis Method:", metadata.get('extraction_method', 'Legal-BERT AI')),
            ("System Version:", metadata.get('system_version', 'AI Legal Timeline Builder Pro'))
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Statistics section
        row += 2
        ws[f'A{row}'] = "Timeline Statistics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        # Calculate statistics
        total_events = len(timeline_events)
        high_confidence = len([e for e in timeline_events if e.get('confidence', 0) > 0.8])
        medium_confidence = len([e for e in timeline_events if 0.5 < e.get('confidence', 0) <= 0.8])
        low_confidence = len([e for e in timeline_events if e.get('confidence', 0) <= 0.5])

        stats_data = [
            ("Total Events", total_events),
            ("High Confidence (>80%)", high_confidence),
            ("Medium Confidence (50-80%)", medium_confidence),
            ("Low Confidence (<50%)", low_confidence),
            ("Average Confidence", f"{sum(e.get('confidence', 0) for e in timeline_events) / total_events:.1%}" if total_events > 0 else "N/A")
        ]

        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Source files section
        if metadata.get('source_files'):
            row += 2
            ws[f'A{row}'] = "Source Documents"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1

            for i, filename in enumerate(metadata['source_files'], 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = filename
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def _create_timeline_sheet(self, wb: Workbook, timeline_events: List[Dict[str, Any]]):
        """Create detailed timeline sheet"""
        ws = wb.create_sheet("Timeline")

        # Headers
        headers = ["Date", "Event", "Event Type", "Entities", "Confidence", "Source File", "Text Snippet"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2a5298", end_color="2a5298", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row, event in enumerate(timeline_events, 2):
            ws.cell(row=row, column=1, value=event.get('date', 'Unknown'))
            ws.cell(row=row, column=2, value=event.get('event', ''))
            ws.cell(row=row, column=3, value=event.get('event_type', ''))
            ws.cell(row=row, column=4, value=', '.join(event.get('entities', [])))
            ws.cell(row=row, column=5, value=event.get('confidence', 0))
            ws.cell(row=row, column=6, value=event.get('metadata', {}).get('filename', ''))
            ws.cell(row=row, column=7, value=(event.get('text', '') or '')[:100] + ('...' if len(event.get('text', '') or '') > 100 else ''))

            # Format confidence as percentage
            ws.cell(row=row, column=5).number_format = '0.0%'

            # Color code by confidence
            confidence = event.get('confidence', 0)
            if confidence > 0.8:
                fill_color = "d4edda"  # Green
            elif confidence > 0.5:
                fill_color = "fff3cd"  # Yellow
            else:
                fill_color = "f8d7da"  # Red

            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(timeline_events)+1, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border

    def _create_analysis_sheet(self, wb: Workbook, timeline_events: List[Dict[str, Any]]):
        """Create analysis sheet with pivot tables and charts"""
        ws = wb.create_sheet("Analysis")

        # Event type analysis
        ws['A1'] = "Event Type Analysis"
        ws['A1'].font = Font(size=14, bold=True)

        # Count events by type
        event_types = {}
        for event in timeline_events:
            event_type = event.get('event_type', 'Unknown')
            event_types[event_type] = event_types.get(event_type, 0) + 1

        # Create table
        ws['A3'] = "Event Type"
        ws['B3'] = "Count"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)

        row = 4
        for event_type, count in sorted(event_types.items(), key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = event_type
            ws[f'B{row}'] = count
            row += 1

        # Confidence distribution
        ws[f'A{row + 2}'] = "Confidence Distribution"
        ws[f'A{row + 2}'].font = Font(size=14, bold=True)

        confidence_ranges = {
            "High (>80%)": len([e for e in timeline_events if e.get('confidence', 0) > 0.8]),
            "Medium (50-80%)": len([e for e in timeline_events if 0.5 < e.get('confidence', 0) <= 0.8]),
            "Low (<50%)": len([e for e in timeline_events if e.get('confidence', 0) <= 0.5])
        }

        row += 4
        ws[f'A{row}'] = "Confidence Range"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        for conf_range, count in confidence_ranges.items():
            ws[f'A{row}'] = conf_range
            ws[f'B{row}'] = count
            row += 1

    def _create_evidence_sheet(self, wb: Workbook, timeline_events: List[Dict[str, Any]]):
        """Create evidence tracking sheet"""
        ws = wb.create_sheet("Evidence")

        # Headers
        headers = ["Event ID", "Date", "Event", "Source File", "Confidence", "Evidence Citation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2a5298", end_color="2a5298", fill_type="solid")

        # Data rows
        for row, event in enumerate(timeline_events, 2):
            event_id = f"EVT-{row-1:03d}"
            source_file = event.get('metadata', {}).get('filename', 'Unknown')
            confidence = event.get('confidence', 0)

            ws.cell(row=row, column=1, value=event_id)
            ws.cell(row=row, column=2, value=event.get('date', 'Unknown'))
            ws.cell(row=row, column=3, value=event.get('event', ''))
            ws.cell(row=row, column=4, value=source_file)
            ws.cell(row=row, column=5, value=confidence)
            ws.cell(row=row, column=6, value=f"[Source: {source_file}] (Confidence: {confidence:.1%})")

            # Format confidence as percentage
            ws.cell(row=row, column=5).number_format = '0.0%'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
